import json
from openpyxl import load_workbook, Workbook
import os

histDir = r'./controller/history.json'

ctrXlDir = r'./controller/clients.xlsx'

dbXlDir = r'./db/data.xlsx'


def readUrl(key):
    wb = load_workbook(ctrXlDir)
    
    ws = wb.active
    
    return [ws[f'A{key + 1}'].value, ws[f'B{key + 1}'].value]
    
def fetchJobs():
    wb = load_workbook(ctrXlDir)
    
    ws = wb.active
    
    return [cell.value.lower() for cell in ws['D']]


def updateDB(arr):
    if os.path.exists(dbXlDir):
        wb = load_workbook(dbXlDir)
    else:
        wb = Workbook()

    if wb.active:
        ws = wb.active
    else:
        ws = wb.create_sheet()
        
    if ws["A1"] != "Job Title":
        ws["A1"] = "Job Title"
        ws["B1"] = "Company"
        ws["C1"] = "Location"
        ws["D1"] = "Url"
    
    for item in arr:
        for job in fetchJobs():
            if (job in item[0].lower()):
                ws.append(item)
                break

    wb.save(dbXlDir)
    

def readHistory(key=None):
    with open(histDir, 'r') as file:
        try:
            data = json.load(file)
            
            if key is not None:
                return data.get(f'{key}', []) 
            
            return data
        except:
            return []
        
        
def updateHistory(key, val):
    data = readHistory()
    data[key] = val
    
    with open(histDir, 'w') as file:
        try:
            json.dump(data, file, indent=4)
        except Exception as e:
            print(e)